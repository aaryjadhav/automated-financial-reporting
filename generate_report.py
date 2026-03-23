import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import time

# Colour palette
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
BLACK      = "000000"

# ── STEP 1: PULL DATA ──────────────────────────────────────
def pull_data():
    print("Pulling raw data...")
    sales_df    = pd.read_csv("data/raw_sales.csv")
    expenses_df = pd.read_csv("data/raw_expenses.csv")
    print(f"   Sales rows    : {len(sales_df)}")
    print(f"   Expenses rows : {len(expenses_df)}")
    return sales_df, expenses_df

# ── STEP 2: CLEAN DATA ─────────────────────────────────────
def clean_currency(series):
    return (
        series.astype(str)
              .str.replace(r"[\$,\s]", "", regex=True)
              .replace("", np.nan)
              .astype(float)
    )

def clean_sales(df):
    print("Cleaning sales data...")
    df = df.copy()

    # Convert "Jan-2024" text into a real date Python understands
    df["Month_dt"]    = pd.to_datetime(df["Month"], format="%b-%Y")
    df["Month_Label"] = df["Month_dt"].dt.strftime("%b %Y")

    # Clean the messy columns
    df["Revenue"]    = clean_currency(df["Revenue"])
    df["Units_Sold"] = pd.to_numeric(df["Units_Sold"], errors="coerce")
    df["Returns"]    = pd.to_numeric(df["Returns"],    errors="coerce").fillna(0)

    # Create new calculated columns
    df["Net_Revenue"] = df["Revenue"] - (df["Returns"] / df["Units_Sold"] * df["Revenue"])
    df["Return_Rate"] = df["Returns"] / df["Units_Sold"]

    # Remove rows where critical data is missing
    before = len(df)
    df.dropna(subset=["Revenue", "Units_Sold"], inplace=True)
    print(f"   Dropped {before - len(df)} rows with missing data")
    print(f"   Sales clean: {len(df)} rows remaining")

    return df.sort_values("Month_dt").reset_index(drop=True)

def clean_expenses(df):
    print("Cleaning expenses data...")
    df = df.copy()

    df["Month_dt"]    = pd.to_datetime(df["Month"], format="%b-%Y")
    df["Month_Label"] = df["Month_dt"].dt.strftime("%b %Y")
    df["Amount"]      = clean_currency(df["Amount"])
    df["Category"]    = df["Category"].str.strip().str.title()

    before = len(df)
    df.dropna(subset=["Amount"], inplace=True)
    print(f"   Dropped {before - len(df)} rows with missing Amount")
    print(f"   Expenses clean: {len(df)} rows remaining")

    return df.sort_values("Month_dt").reset_index(drop=True)

# ── STEP 3: BUILD SUMMARY TABLES ───────────────────────────

def build_monthly_pnl(sales_df, expenses_df):
    # Total revenue per month from sales data
    rev = (
        sales_df.groupby("Month_dt")
                .agg(
                    Revenue    = ("Revenue",     "sum"),
                    Net_Revenue= ("Net_Revenue", "sum"),
                    Units_Sold = ("Units_Sold",  "sum")
                )
                .reset_index()
    )

    # Total expenses per month
    exp = (
        expenses_df.groupby("Month_dt")
                   .agg(Total_Expenses=("Amount", "sum"))
                   .reset_index()
    )

    # COGS only (Cost of Goods Sold) — just the manufacturing rows
    cogs = (
        expenses_df[expenses_df["Category"] == "Cogs"]
                   .groupby("Month_dt")
                   .agg(COGS=("Amount", "sum"))
                   .reset_index()
    )

    # Merge all three tables together on the Month column
    pnl = rev.merge(exp,  on="Month_dt", how="left") \
             .merge(cogs, on="Month_dt", how="left")

    # Calculate the key financial metrics
    pnl["COGS"]         = pnl["COGS"].fillna(0)
    pnl["Gross_Profit"] = pnl["Revenue"] - pnl["COGS"]
    pnl["Gross_Margin"] = pnl["Gross_Profit"] / pnl["Revenue"]
    pnl["Net_Income"]   = pnl["Revenue"] - pnl["Total_Expenses"]
    pnl["Net_Margin"]   = pnl["Net_Income"] / pnl["Revenue"]
    pnl["Month_Label"]  = pnl["Month_dt"].dt.strftime("%b %Y")

    return pnl.sort_values("Month_dt").reset_index(drop=True)

def build_regional_summary(sales_df):
    return (
        sales_df.groupby("Region")
                .agg(
                    Revenue    = ("Revenue",     "sum"),
                    Units_Sold = ("Units_Sold",  "sum"),
                    Net_Revenue= ("Net_Revenue", "sum"),
                    Returns    = ("Returns",     "sum")
                )
                .reset_index()
                .sort_values("Revenue", ascending=False)
    )


def build_product_summary(sales_df):
    return (
        sales_df.groupby("Product")
                .agg(
                    Revenue    = ("Revenue",     "sum"),
                    Units_Sold = ("Units_Sold",  "sum"),
                    Net_Revenue= ("Net_Revenue", "sum")
                )
                .reset_index()
                .sort_values("Revenue", ascending=False)
    )


def build_expense_breakdown(expenses_df):
    return (
        expenses_df.groupby("Category")
                   .agg(Total=("Amount", "sum"))
                   .reset_index()
                   .sort_values("Total", ascending=False)
    )

# ── STEP 4: STYLE HELPERS ──────────────────────────────────

def hdr_font(size=11, color=WHITE):
    return Font(name="Arial", size=size, bold=True, color=color)

def body_font(size=10, bold=False, color=BLACK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def left_align():
    return Alignment(horizontal="left", vertical="center")

def apply_header_row(ws, row, col_start, col_end, labels):
    for i, label in enumerate(labels, start=col_start):
        c = ws.cell(row=row, column=i, value=label)
        c.font      = hdr_font()
        c.fill      = fill(MID_BLUE)
        c.alignment = center()
        c.border    = thin_border()

def apply_data_row(ws, row, col_start, values, shade=False):
    bg = LIGHT_GRAY if shade else WHITE
    for i, val in enumerate(values, start=col_start):
        c = ws.cell(row=row, column=i, value=val)
        c.font      = body_font()
        c.fill      = fill(bg)
        c.border    = thin_border()
        c.alignment = right() if isinstance(val, (int, float)) else left_align()
        if isinstance(val, float) and abs(val) < 1 and val != 0:
            c.number_format = "0.0%"
        elif isinstance(val, (int, float)):
            c.number_format = '#,##0;-#,##0;"-"'

# ── STEP 5: WRITE EXCEL SHEETS ─────────────────────────────

def write_pnl(ws, pnl):
    ws.sheet_view.showGridLines = False

    # Set column widths
    for col, width in zip("BCDEFGH", [14, 16, 16, 16, 16, 13, 13]):
        ws.column_dimensions[col].width = width

    # Title banner
    ws.merge_cells("B1:H2")
    c = ws["B1"]
    c.value     = "MONTHLY PROFIT & LOSS — FY 2024"
    c.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = center()

    # Header row
    headers = ["Month", "Revenue ($)", "COGS ($)", "Gross Profit ($)",
               "Total Expenses ($)", "Net Income ($)", "Gross Margin", "Net Margin"]
    apply_header_row(ws, 4, 2, 9, headers)

    # Data rows
    for i, row_data in pnl.iterrows():
        r = i + 5
        vals = [
            row_data["Month_Label"],
            row_data["Revenue"],
            row_data["COGS"],
            row_data["Gross_Profit"],
            row_data["Total_Expenses"],
            row_data["Net_Income"],
            row_data["Gross_Margin"],
            row_data["Net_Margin"],
        ]
        apply_data_row(ws, r, 2, vals, shade=(i % 2 == 0))

    # Totals row
    tot_r = len(pnl) + 5
    totals = ["FULL YEAR",
              pnl["Revenue"].sum(), pnl["COGS"].sum(),
              pnl["Gross_Profit"].sum(), pnl["Total_Expenses"].sum(),
              pnl["Net_Income"].sum(),
              pnl["Gross_Margin"].mean(), pnl["Net_Margin"].mean()]
    for i, val in enumerate(totals, start=2):
        c = ws.cell(row=tot_r, column=i, value=val)
        c.font      = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill      = fill(DARK_BLUE)
        c.border    = thin_border()
        c.alignment = right() if isinstance(val, (int, float)) else center()
        if isinstance(val, float) and abs(val) < 1:
            c.number_format = "0.0%"
        elif isinstance(val, (int, float)):
            c.number_format = '#,##0;-#,##0;"-"'

    # Line chart — Revenue vs Net Income
    chart = LineChart()
    chart.title  = "Monthly Revenue vs Net Income"
    chart.style  = 10
    chart.height = 12
    chart.width  = 24

    data_rev = Reference(ws, min_col=3, min_row=4, max_row=tot_r - 1)
    data_ni  = Reference(ws, min_col=7, min_row=4, max_row=tot_r - 1)
    cats     = Reference(ws, min_col=2, min_row=5, max_row=tot_r - 1)

    chart.add_data(data_rev, titles_from_data=True)
    chart.add_data(data_ni,  titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, f"B{tot_r + 2}")

def write_regional(ws, regional):
    ws.sheet_view.showGridLines = False
    for col, width in zip("BCDEF", [16, 18, 14, 18, 13]):
        ws.column_dimensions[col].width = width

    ws.merge_cells("B1:F2")
    c = ws["B1"]
    c.value     = "REGIONAL SALES BREAKDOWN — FY 2024"
    c.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = center()

    headers = ["Region", "Revenue ($)", "Units Sold", "Net Revenue ($)", "Returns"]
    apply_header_row(ws, 4, 2, 6, headers)

    for i, row_data in regional.iterrows():
        vals = [row_data["Region"], row_data["Revenue"],
                row_data["Units_Sold"], row_data["Net_Revenue"],
                row_data["Returns"]]
        apply_data_row(ws, i + 5, 2, vals, shade=(i % 2 == 0))

    # Bar chart
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "Revenue by Region"
    chart.style  = 10
    chart.height = 12
    chart.width  = 20
    data = Reference(ws, min_col=3, min_row=4, max_row=4 + len(regional))
    cats = Reference(ws, min_col=2, min_row=5, max_row=4 + len(regional))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "B10")   

def write_expenses(ws, exp_breakdown):
    ws.sheet_view.showGridLines = False
    for col, width in zip("BCD", [20, 18, 12]):
        ws.column_dimensions[col].width = width

    ws.merge_cells("B1:D2")
    c = ws["B1"]
    c.value     = "EXPENSE BREAKDOWN — FY 2024"
    c.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = center()

    headers = ["Category", "Annual Total ($)", "% of Total"]
    apply_header_row(ws, 4, 2, 4, headers)

    grand_total = exp_breakdown["Total"].sum()
    for i, row_data in exp_breakdown.iterrows():
        pct  = row_data["Total"] / grand_total
        vals = [row_data["Category"], row_data["Total"], pct]
        apply_data_row(ws, i + 5, 2, vals, shade=(i % 2 == 0))

    # Grand total row
    tot_r = len(exp_breakdown) + 5
    for col, val in zip([2, 3, 4], ["GRAND TOTAL", grand_total, 1.0]):
        c = ws.cell(row=tot_r, column=col, value=val)
        c.font   = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill   = fill(DARK_BLUE)
        c.border = thin_border()
        c.alignment = right() if isinstance(val, (int, float)) else center()
        if isinstance(val, float) and abs(val) <= 1:
            c.number_format = "0.0%"
        elif isinstance(val, (int, float)):
            c.number_format = '#,##0;-#,##0;"-"'



# ── MAIN ───────────────────────────────────────────────────
def main():
    start_time = time.time()

    print("\n" + "="*45)
    print("   AUTOMATED FINANCIAL REPORTING SYSTEM")
    print("="*45)

    # Pull and clean
    sales_raw, expenses_raw = pull_data()
    sales_df    = clean_sales(sales_raw)
    expenses_df = clean_expenses(expenses_raw)

    # Build summaries
    print("\nBuilding summary tables...")
    pnl      = build_monthly_pnl(sales_df, expenses_df)
    regional = build_regional_summary(sales_df)
    product  = build_product_summary(sales_df)
    exp_break= build_expense_breakdown(expenses_df)

    # Build Excel workbook
    print("\nGenerating Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    # Create each sheet and write to it
    sheets = [
        ("Monthly P&L",        lambda ws: write_pnl(ws, pnl)),
        ("Regional Breakdown", lambda ws: write_regional(ws, regional)),
        ("Expense Breakdown",  lambda ws: write_expenses(ws, exp_break)),
    ]

    for name, writer_fn in sheets:
        ws = wb.create_sheet(name)
        writer_fn(ws)
        print(f"   Sheet written: {name}")

    # Colour the tabs at the bottom
    wb["Monthly P&L"].sheet_properties.tabColor        = "1A7A4A"
    wb["Regional Breakdown"].sheet_properties.tabColor = "8E44AD"
    wb["Expense Breakdown"].sheet_properties.tabColor  = "C0392B"

    wb.save("Financial_Report_2024.xlsx")
    elapsed = round(time.time() - start_time, 2)
    print(f"\n  Report saved  →  Financial_Report_2024.xlsx")
    print(f"  Time taken    →  {elapsed} seconds")
    print("="*45 + "\n")

if __name__ == "__main__":
    main()